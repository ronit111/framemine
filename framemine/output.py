"""Output formatters: JSON and Excel."""

import json
import logging
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

logger = logging.getLogger(__name__)


def write_json(items: list[dict], output_path: Path, indent: int = 2) -> Path:
    """Write items to JSON. Returns path."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(items, f, indent=indent, ensure_ascii=False)
    logger.info("Wrote %d items to %s", len(items), output_path)
    return output_path


def write_excel(
    items: list[dict],
    output_path: Path,
    columns: list[str] | None = None,
    sheet_name: str = "Extracted Data",
) -> Path:
    """
    Write formatted Excel:
    - Dark header (#2D2D2D) with white bold text
    - Auto-column-width (capped at 60)
    - Frozen header row
    - Auto-filter
    - Hyperlinked source URLs (blue, underlined)
    - Row numbers in first column

    If columns is None, infer from first item's keys.
    Returns path.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if columns is None:
        if items:
            columns = list(items[0].keys())
        else:
            columns = []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header row: "#" + user-supplied columns
    all_headers = ["#"] + columns
    header_fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")

    for col_idx, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_num, item in enumerate(items, 1):
        ws.cell(row=row_num + 1, column=1, value=row_num)
        for col_idx, col_name in enumerate(columns, 2):
            value = item.get(col_name, "")
            if value is None:
                value = ""
            cell = ws.cell(row=row_num + 1, column=col_idx, value=value)

            # Hyperlink source URLs
            if col_name in ("source", "source_url", "url") and isinstance(value, str) and value.startswith("http"):
                cell.hyperlink = value
                cell.font = Font(color="4A86C8", underline="single")

    # Auto-column-width (capped at 60)
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    if ws.dimensions:
        ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    logger.info("Wrote %d items to %s", len(items), output_path)
    return output_path


def write_outputs(
    items: list[dict],
    output_path: Path,
    columns: list[str] | None = None,
    sheet_name: str = "Extracted Data",
    json_output: bool = True,
    excel_output: bool = True,
) -> dict[str, Path]:
    """Write all requested formats. Returns {"json": Path, "excel": Path}."""
    output_path = Path(output_path)
    results: dict[str, Path] = {}

    if json_output:
        json_path = output_path.with_suffix(".json")
        results["json"] = write_json(items, json_path)

    if excel_output:
        excel_path = output_path.with_suffix(".xlsx")
        results["excel"] = write_excel(items, excel_path, columns=columns, sheet_name=sheet_name)

    return results
