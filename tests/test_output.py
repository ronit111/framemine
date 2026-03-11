"""Tests for framemine.output module."""

import json

import openpyxl

from framemine.output import write_excel, write_json, write_outputs


SAMPLE_ITEMS = [
    {"title": "The Great Gatsby", "author": "F. Scott Fitzgerald", "source": "https://example.com/1"},
    {"title": "Sapiens", "author": "Yuval Noah Harari", "source": "https://example.com/2"},
]


class TestWriteJson:
    def test_write_json_creates_valid_file(self, tmp_path):
        path = tmp_path / "output.json"
        result = write_json(SAMPLE_ITEMS, path)

        assert result == path
        assert path.exists()

        with open(path) as f:
            loaded = json.load(f)
        assert loaded == SAMPLE_ITEMS

    def test_write_json_empty_list(self, tmp_path):
        path = tmp_path / "empty.json"
        write_json([], path)

        with open(path) as f:
            content = f.read().strip()
        assert content == "[]"


class TestWriteExcel:
    def test_write_excel_creates_file(self, tmp_path):
        path = tmp_path / "output.xlsx"
        result = write_excel(SAMPLE_ITEMS, path)

        assert result == path
        assert path.exists()

        # Verify it can be opened
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.max_row == 3  # header + 2 data rows
        wb.close()

    def test_write_excel_header_formatting(self, tmp_path):
        path = tmp_path / "output.xlsx"
        write_excel(SAMPLE_ITEMS, path)

        wb = openpyxl.load_workbook(path)
        ws = wb.active

        header_cell = ws.cell(row=1, column=1)
        assert header_cell.fill.start_color.rgb == "002D2D2D"
        assert header_cell.font.bold is True
        assert header_cell.font.color.rgb == "00FFFFFF"
        wb.close()

    def test_write_excel_hyperlinks(self, tmp_path):
        path = tmp_path / "output.xlsx"
        write_excel(SAMPLE_ITEMS, path, columns=["title", "author", "source"])

        wb = openpyxl.load_workbook(path)
        ws = wb.active

        # source is column 4 (1=#, 2=title, 3=author, 4=source)
        source_cell = ws.cell(row=2, column=4)
        assert source_cell.hyperlink is not None
        assert source_cell.hyperlink.target == "https://example.com/1"
        wb.close()

    def test_write_excel_frozen_panes(self, tmp_path):
        path = tmp_path / "output.xlsx"
        write_excel(SAMPLE_ITEMS, path)

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.freeze_panes == "A2"
        wb.close()

    def test_write_excel_infers_columns(self, tmp_path):
        path = tmp_path / "output.xlsx"
        write_excel(SAMPLE_ITEMS, path)  # columns=None

        wb = openpyxl.load_workbook(path)
        ws = wb.active

        # Headers: #, title, author, source (inferred from first item keys)
        headers = [ws.cell(row=1, column=col).value for col in range(1, 5)]
        assert headers == ["#", "title", "author", "source"]
        wb.close()


class TestWriteOutputs:
    def test_write_outputs_creates_both(self, tmp_path):
        path = tmp_path / "result"
        results = write_outputs(
            SAMPLE_ITEMS, path,
            json_output=True, excel_output=True,
        )

        assert "json" in results
        assert "excel" in results
        assert results["json"].suffix == ".json"
        assert results["excel"].suffix == ".xlsx"
        assert results["json"].exists()
        assert results["excel"].exists()

    def test_write_outputs_json_only(self, tmp_path):
        path = tmp_path / "result"
        results = write_outputs(
            SAMPLE_ITEMS, path,
            json_output=True, excel_output=False,
        )

        assert "json" in results
        assert "excel" not in results
        assert results["json"].exists()
